# -*- coding: utf-8 -*-
import openpyxl, glob, os

# 找到最新上传的文体Excel文件
excel_dir = r"d:\wp\waibao\gongcheng\backend\uploads\excel"
wenti_dir = r"d:\wp\waibao\gongcheng\文体"

# 检查文体目录下的xlsx文件
for d in [wenti_dir, excel_dir]:
    if not os.path.exists(d):
        continue
    for f in os.listdir(d):
        if f.endswith(".xlsx"):
            path = os.path.join(d, f)
            print(f"\n=== File: {f} ===")
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            print(f"Sheet: {ws.title}, rows: {ws.max_row}, cols: {ws.max_column}")

            # Print headers (row 1 and row 2)
            for r in [1, 2]:
                vals = []
                for c in range(1, min(ws.max_column + 1, 20)):
                    v = ws.cell(r, c).value
                    vals.append(str(v)[:20] if v else "")
                print(f"  Row {r}: {vals}")

            # Print first data row
            if ws.max_row >= 3:
                vals = []
                for c in range(1, min(ws.max_column + 1, 20)):
                    v = ws.cell(3, c).value
                    vals.append(str(v)[:30] if v else "")
                print(f"  Row 3 (data): {vals}")

            wb.close()
