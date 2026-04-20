# -*- coding: utf-8 -*-
import openpyxl

path = r"d:\wp\waibao\gongcheng\文体\隐患依据(2).xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\nSheet: '{name}', rows: {ws.max_row}, cols: {ws.max_column}")
    # Print first 3 rows to see structure
    for r in range(1, min(5, ws.max_row + 1)):
        row_data = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            row_data.append(str(v)[:40] if v else "")
        print(f"  Row {r}: {row_data}")

wb.close()
