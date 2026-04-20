# -*- coding: utf-8 -*-
"""
Excel解析引擎：解析分段式安全隐患问题清单
- 按街道分组，每个项目重复"项目信息行 -> 表头行 -> 数据行"
- 提取嵌入图片（TwoCellAnchor + DISPIMG 两种方式）
"""
import re
import os
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import List, Dict, Optional
from dataclasses import dataclass, field

import openpyxl
from PIL import Image
import io


@dataclass
class HazardData:
    seq: int
    hazard_type: str
    description: str
    risk: str
    category: str
    reference: str = ""
    remark: str = ""
    hazard_photo_path: Optional[str] = None


@dataclass
class ProjectData:
    name: str
    street: str
    address: str = ""
    contact: str = ""
    phone: str = ""
    category: str = "建筑施工"
    build_unit: str = ""
    construct_unit: str = ""
    supervise_unit: str = ""
    check_date: str = ""
    seq_in_street: int = 0
    hazards: List[HazardData] = field(default_factory=list)


def _clean(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _parse_contact_cell(c4_text: str, info: dict):
    """Parse address / contact / phone from C4 cell into info dict."""
    if not c4_text:
        return
    lines = c4_text.replace("\r\n", "\n").replace("\r", "\n")
    m = re.search(r"地\s*址[：:]\s*(.+?)(?:\n|$)", lines)
    if m:
        info["address"] = m.group(1).strip()
    m = re.search(r"联系人[：:]\s*(.+?)(?:\n|$)", lines)
    if m:
        info["contact"] = m.group(1).strip()
    m = re.search(r"联系电话[：:]\s*(.+?)(?:\n|$)", lines)
    if m:
        info["phone"] = m.group(1).strip()


def _parse_project_info(c1_text: str, c4_text: str) -> dict:
    """Parse project info from the merged cells C1 and C4 (建筑施工 format)."""
    info = {
        "name": "", "build_unit": "", "construct_unit": "",
        "supervise_unit": "", "address": "", "contact": "", "phone": ""
    }
    if c1_text:
        lines = c1_text.replace("\r\n", "\n").replace("\r", "\n")
        m = re.search(r"项目名称[：:]\s*(.+?)(?:\n|$)", lines)
        if m:
            info["name"] = m.group(1).strip()
        m = re.search(r"建设单位[：:]\s*(.+?)(?:\n|$)", lines)
        if m:
            info["build_unit"] = m.group(1).strip()
        m = re.search(r"施工单位[：:]\s*(.+?)(?:\n|$)", lines)
        if m:
            info["construct_unit"] = m.group(1).strip()
        m = re.search(r"监理单位[：:]\s*(.+?)(?:\n|$)", lines)
        if m:
            info["supervise_unit"] = m.group(1).strip()

    _parse_contact_cell(c4_text, info)
    return info


_RE_NUMBERED_PROJECT = re.compile(r"^\d+[、,.．]\s*(.+)")


def _parse_simple_project_name(c1_text: str) -> Optional[str]:
    """Try to parse 'N、企业名称' format. Returns name or None."""
    m = _RE_NUMBERED_PROJECT.match(c1_text)
    return m.group(1).strip() if m else None


def _extract_images_from_xlsx(xlsx_path: str, output_dir: str) -> Dict[str, str]:
    """
    Extract all images from xlsx and return mapping: "row_col" -> saved_path.
    Row and col are 0-based anchor positions.
    Uses openpyxl ws._images first, then falls back to direct drawing XML parsing.
    """
    os.makedirs(output_dir, exist_ok=True)
    image_map: Dict[str, str] = {}

    image_map = _extract_via_openpyxl(xlsx_path, output_dir)

    xml_map = _extract_via_drawing_xml(xlsx_path, output_dir)
    for key, path in xml_map.items():
        if key not in image_map:
            image_map[key] = path

    return image_map


def _extract_via_openpyxl(xlsx_path: str, output_dir: str) -> Dict[str, str]:
    """Primary extraction using openpyxl's image detection."""
    image_map: Dict[str, str] = {}
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        for idx, img in enumerate(ws._images):
            anchor = img.anchor
            from_marker = getattr(anchor, '_from', None)
            if from_marker is None:
                continue
            row = from_marker.row
            col = from_marker.col

            ext = "png"
            img_data = img._data()
            if img_data[:3] == b'\xff\xd8\xff':
                ext = "jpg"

            filename = f"img_r{row}_c{col}_{idx}.{ext}"
            filepath = os.path.join(output_dir, filename)
            with open(filepath, "wb") as f:
                f.write(img_data)

            key = f"{row}_{col}"
            image_map[key] = filepath
        wb.close()
    except Exception:
        pass
    return image_map


XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _extract_via_drawing_xml(xlsx_path: str, output_dir: str) -> Dict[str, str]:
    """
    Fallback: parse xl/drawings/drawing*.xml directly to build cell-to-image mapping.
    Handles cases openpyxl misses (OneCellAnchor, shifted anchors, etc.).
    """
    image_map: Dict[str, str] = {}
    try:
        with zipfile.ZipFile(xlsx_path) as zf:
            drawing_names = [
                n for n in zf.namelist()
                if n.startswith("xl/drawings/drawing") and n.endswith(".xml")
            ]
            for drawing_path in drawing_names:
                rels_path = drawing_path.replace(
                    "xl/drawings/", "xl/drawings/_rels/"
                ) + ".rels"
                rels_map = _parse_rels(zf, rels_path)
                drawing_xml = zf.read(drawing_path).decode("utf-8")
                root = ET.fromstring(drawing_xml)

                for anchor_el in root:
                    tag = anchor_el.tag.split("}")[-1] if "}" in anchor_el.tag else anchor_el.tag
                    if tag not in ("twoCellAnchor", "oneCellAnchor"):
                        continue

                    from_el = anchor_el.find(f"{{{XDR_NS}}}from")
                    if from_el is None:
                        continue
                    row_el = from_el.find(f"{{{XDR_NS}}}row")
                    col_el = from_el.find(f"{{{XDR_NS}}}col")
                    if row_el is None or col_el is None:
                        continue
                    row = int(row_el.text)
                    col = int(col_el.text)

                    embed_id = None
                    for blip in anchor_el.iter(f"{{{A_NS}}}blip"):
                        embed_id = blip.get(f"{{{R_NS}}}embed")
                        if embed_id:
                            break
                    if not embed_id or embed_id not in rels_map:
                        continue

                    media_zip_path = rels_map[embed_id]
                    if not media_zip_path.startswith("xl/"):
                        media_zip_path = "xl/" + media_zip_path.lstrip("../")
                    if media_zip_path not in zf.namelist():
                        alt = media_zip_path.replace("xl/xl/", "xl/")
                        if alt in zf.namelist():
                            media_zip_path = alt
                        else:
                            continue

                    data = zf.read(media_zip_path)
                    ext = os.path.splitext(media_zip_path)[1].lstrip(".") or "png"
                    if ext not in ("png", "jpg", "jpeg", "gif", "bmp", "webp"):
                        ext = "png"
                    filename = f"xml_r{row}_c{col}.{ext}"
                    filepath = os.path.join(output_dir, filename)
                    if not os.path.exists(filepath):
                        with open(filepath, "wb") as f:
                            f.write(data)

                    key = f"{row}_{col}"
                    if key not in image_map:
                        image_map[key] = filepath

    except Exception:
        pass
    return image_map


def _parse_rels(zf: zipfile.ZipFile, rels_path: str) -> Dict[str, str]:
    """Parse a .rels file and return {rId -> target_path}."""
    rels_map: Dict[str, str] = {}
    if rels_path not in zf.namelist():
        return rels_map
    try:
        rels_xml = zf.read(rels_path).decode("utf-8")
        rels_root = ET.fromstring(rels_xml)
        for rel in rels_root:
            rid = rel.get("Id", "")
            target = rel.get("Target", "")
            if target.startswith("../"):
                target = "xl/" + target[3:]
            rels_map[rid] = target
    except Exception:
        pass
    return rels_map


PHOTO_COL_CANDIDATES = [4, 3, 5]


def _find_photo_for_row(image_map: Dict[str, str], anchor_row: int) -> Optional[str]:
    """Try column 4 first (expected photo column), then 3 and 5 as fallback."""
    for col in PHOTO_COL_CANDIDATES:
        key = f"{anchor_row}_{col}"
        path = image_map.get(key)
        if path and os.path.exists(path):
            return path
    return None


def _detect_category_from_title(ws, max_scan: int = 5) -> str:
    """Scan the first few rows for a title containing category hints."""
    for r in range(1, min(max_scan + 1, ws.max_row + 1)):
        val = _clean(ws.cell(r, 1).value)
        if "九小场所" in val:
            return "九小场所"
        if "建筑施工" in val:
            return "建筑施工"
    return ""


def _read_hazard_rows(ws, start_row: int, max_row: int, image_map: Dict[str, str]) -> tuple:
    """Read consecutive hazard data rows starting from start_row.
    Returns (list_of_HazardData, next_row)."""
    hazards: List[HazardData] = []
    row = start_row
    while row <= max_row:
        seq_val = ws.cell(row, 1).value
        if seq_val is None:
            row += 1
            continue

        seq_str = _clean(seq_val)
        if not seq_str.isdigit():
            break

        hz = HazardData(
            seq=int(seq_str),
            hazard_type=_clean(ws.cell(row, 2).value),
            description=_clean(ws.cell(row, 3).value),
            risk=_clean(ws.cell(row, 4).value),
            category=_clean(ws.cell(row, 6).value),
            reference=_clean(ws.cell(row, 7).value),
            remark=_clean(ws.cell(row, 8).value),
        )

        anchor_row = row - 1  # 0-based
        photo_path = _find_photo_for_row(image_map, anchor_row)
        if photo_path:
            hz.hazard_photo_path = photo_path

        hazards.append(hz)
        row += 1

    return hazards, row


def parse_excel(xlsx_path: str, photo_output_dir: str) -> List[ProjectData]:
    """
    Main entry: parse the safety inspection Excel file.
    Supports two formats:
      1) 建筑施工 format: "项目名称：xxx" in C1
      2) 九小场所 format: "N、企业名称" in C1
    Returns a list of ProjectData with hazards and photo paths.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    max_row = ws.max_row

    image_map = _extract_images_from_xlsx(xlsx_path, photo_output_dir)

    file_category = _detect_category_from_title(ws)

    projects: List[ProjectData] = []
    current_street = ""
    street_seq = 0

    row = 1
    while row <= max_row:
        c1 = _clean(ws.cell(row, 1).value)

        if not c1:
            row += 1
            continue

        if c1 in ("附件",) or "专项检查安全隐患" in c1:
            row += 1
            continue

        if re.match(r"^[\u4e00-\u9fa5]{2,6}街道$", c1):
            current_street = c1
            street_seq = 0
            row += 1
            continue

        # Format 1: 建筑施工 — "项目名称：xxx"
        if re.search(r"项目名称[：:]", c1):
            street_seq += 1
            c4 = _clean(ws.cell(row, 4).value)
            info = _parse_project_info(c1, c4)

            proj = ProjectData(
                name=info["name"],
                street=current_street,
                address=info["address"],
                contact=info["contact"],
                phone=info["phone"],
                category=file_category or "建筑施工",
                build_unit=info["build_unit"],
                construct_unit=info["construct_unit"],
                supervise_unit=info["supervise_unit"],
                seq_in_street=street_seq,
            )

            row += 1
            if row <= max_row and _clean(ws.cell(row, 1).value) == "序号":
                row += 1

            hazards, row = _read_hazard_rows(ws, row, max_row, image_map)
            proj.hazards = hazards
            projects.append(proj)
            continue

        # Format 2: 九小场所 — "N、企业名称"
        simple_name = _parse_simple_project_name(c1)
        if simple_name:
            street_seq += 1
            c4 = _clean(ws.cell(row, 4).value)
            info = {
                "name": simple_name, "address": "", "contact": "", "phone": "",
                "build_unit": "", "construct_unit": "", "supervise_unit": "",
            }
            _parse_contact_cell(c4, info)

            proj = ProjectData(
                name=info["name"],
                street=current_street,
                address=info["address"],
                contact=info["contact"],
                phone=info["phone"],
                category=file_category or "九小场所",
                seq_in_street=street_seq,
            )

            row += 1
            if row <= max_row and _clean(ws.cell(row, 1).value) == "序号":
                row += 1

            hazards, row = _read_hazard_rows(ws, row, max_row, image_map)
            proj.hazards = hazards
            projects.append(proj)
            continue

        row += 1

    wb.close()
    return projects


if __name__ == "__main__":
    import sys, io as _io
    sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    xlsx = r"d:\wp\gongcheng\doc\附件：3月份建筑施工专项检查安全隐患问题清单.xlsx"
    out = r"d:\wp\gongcheng\backend\uploads\hazard_photos"

    projects = parse_excel(xlsx, out)
    print(f"\n解析完成：共 {len(projects)} 个项目")
    for p in projects:
        photos = sum(1 for h in p.hazards if h.hazard_photo_path)
        print(f"  [{p.street}] {p.name} — 隐患{len(p.hazards)}条, 照片{photos}张")
        for h in p.hazards:
            photo_status = "✓" if h.hazard_photo_path else "✗"
            print(f"    {h.seq}. [{h.hazard_type}] {h.description[:30]}... 照片:{photo_status}")
