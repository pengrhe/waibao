# -*- coding: utf-8 -*-
"""
旅业项目（C包）Excel 解析器：解析旅馆酒店隐患排查台账
列：序号(A) 街道(B) 经营场所名称(C) 工商登记名称(D) 地址(E)
    楼层(F) 面积(G) 现场负责人及联系方式(H) 检查日期(I) 检查人员(J)
    隐患数量(K) 隐患明细(L) 整改建议(M) 引用规范(N) 文件名称(O)
"""
import re
import openpyxl
from typing import List
from dataclasses import dataclass, field


@dataclass
class LvyeHazardData:
    seq: int
    description: str
    suggestion: str
    reference: str = ""


@dataclass
class LvyeProjectData:
    seq: int
    street: str
    venue_name: str
    registered_name: str
    address: str
    floor_info: str
    area: str
    contact: str
    phone: str
    check_date: str
    inspectors: str
    report_code: str
    hazards: List[LvyeHazardData] = field(default_factory=list)


def _clean(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _parse_contact_phone(raw: str):
    """拆分 '姓名\\n手机号' 或 '姓名手机号' 格式。"""
    if not raw:
        return "", ""
    parts = re.split(r'[\n\r]+', raw.strip())
    if len(parts) >= 2:
        return parts[0].strip(), parts[1].strip()
    m = re.match(r'^(.+?)\s*(\d{11})$', raw.strip())
    if m:
        return m.group(1).strip(), m.group(2)
    return raw.strip(), ""


def _parse_numbered_list(text: str) -> List[str]:
    """解析 '1、xxx\\n2、xxx' 格式为列表。"""
    if not text:
        return []
    items = re.split(r'\n\s*(?=\d+[、.])', text.strip())
    result = []
    for item in items:
        cleaned = re.sub(r'^\d+[、.]\s*', '', item.strip())
        if cleaned:
            result.append(cleaned)
    return result


def _extract_report_code(file_name: str) -> str:
    """从文件名称列提取报告编号（如 HC-BA-FHJD-0510）。"""
    if not file_name:
        return ""
    m = re.match(r'(HC-[A-Z]+-[A-Z]+-\d+)', file_name)
    if m:
        return m.group(1)
    return file_name


def _format_date(raw) -> str:
    if not raw:
        return ""
    s = str(raw).strip()
    m = re.match(r'(\d{4})[.\-/年](\d{1,2})[.\-/月](\d{1,2})', s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return s


def parse_lvye_excel(xlsx_path: str) -> List[LvyeProjectData]:
    """
    解析旅业（C包）项目台账 Excel。
    第1行为标题，第2行为表头，第3行起为数据。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    projects: List[LvyeProjectData] = []

    for row in range(3, ws.max_row + 1):
        seq_val = ws.cell(row, 1).value
        if seq_val is None:
            continue
        try:
            seq = int(seq_val)
        except (ValueError, TypeError):
            continue

        venue_name = _clean(ws.cell(row, 3).value)
        if not venue_name:
            continue

        contact_raw = _clean(ws.cell(row, 8).value)
        contact, phone = _parse_contact_phone(contact_raw)

        check_date_raw = ws.cell(row, 9).value
        if hasattr(check_date_raw, 'strftime'):
            check_date = check_date_raw.strftime('%Y-%m-%d')
        else:
            check_date = _format_date(_clean(check_date_raw))

        hazard_text = _clean(ws.cell(row, 12).value)
        suggestion_text = _clean(ws.cell(row, 13).value)
        reference = _clean(ws.cell(row, 14).value)
        file_name = _clean(ws.cell(row, 15).value)

        descriptions = _parse_numbered_list(hazard_text)
        suggestions = _parse_numbered_list(suggestion_text)

        hazards = []
        for i, desc in enumerate(descriptions):
            sug = suggestions[i] if i < len(suggestions) else ""
            hazards.append(LvyeHazardData(
                seq=i + 1,
                description=desc,
                suggestion=sug,
                reference=reference,
            ))

        proj = LvyeProjectData(
            seq=seq,
            street=_clean(ws.cell(row, 2).value),
            venue_name=venue_name,
            registered_name=_clean(ws.cell(row, 4).value),
            address=_clean(ws.cell(row, 5).value),
            floor_info=_clean(ws.cell(row, 6).value),
            area=_clean(ws.cell(row, 7).value),
            contact=contact,
            phone=phone,
            check_date=check_date,
            inspectors=_clean(ws.cell(row, 10).value),
            report_code=_extract_report_code(file_name),
            hazards=hazards,
        )
        projects.append(proj)

    wb.close()
    return projects


if __name__ == "__main__":
    import sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    xlsx = r"d:\wp\waibao\gongcheng\文体\2025年宝安区旅业安全隐患排查(1).xlsx"
    projects = parse_lvye_excel(xlsx)
    print(f"解析完成：共 {len(projects)} 个场所")
    for p in projects[:5]:
        print(f"  [{p.street}] {p.venue_name} ({p.registered_name}) | 编号: {p.report_code}")
        for h in p.hazards:
            print(f"    隐患{h.seq}: {h.description[:40]}...")
            print(f"    建议: {h.suggestion[:40]}...")
