# -*- coding: utf-8 -*-
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

xlsx_path = r'd:\wp\gongcheng\doc\龙华区3月份建筑施工专项检查安全隐患排查清单.xlsx'
wb = openpyxl.load_workbook(xlsx_path, data_only=True)

print('='*80)
print('Excel文件分析')
print('='*80)
print(f'工作表数量: {len(wb.sheetnames)}')
print(f'工作表名称: {wb.sheetnames}')

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n--- 工作表: "{sheet_name}" ---')
    print(f'  行数: {ws.max_row}, 列数: {ws.max_column}')
    
    # 打印前几行
    print('  前10行内容:')
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is not None:
                row_data.append(f'[{col_idx}]{str(val)[:40]}')
        if row_data:
            print(f'    行{row_idx}: {" | ".join(row_data)}')

    # 检查是否有图片
    if hasattr(ws, '_images'):
        print(f'  图片数量: {len(ws._images)}')
        for img in ws._images:
            print(f'    图片位置: {img.anchor}')
