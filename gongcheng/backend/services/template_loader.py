# -*- coding: utf-8 -*-
"""
隐患模板库加载器：解析"隐患依据.xlsx"并导入数据库
"""
import openpyxl
from typing import List, Dict


def _clean(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _detect_start_row(ws) -> int:
    """检测数据起始行：如果第1行A列是数字则无表头从第1行开始，否则从第3行开始。"""
    first_val = ws.cell(1, 1).value
    if first_val is not None:
        try:
            int(first_val)
            return 1
        except (ValueError, TypeError):
            pass
    return 3


def _detect_col_layout(ws, start_row: int) -> bool:
    """检测是否为6列布局（类别+序号+描述+建议+标准+条款）。
    判断依据：第一个数据行的A列不是数字（是文字分类名），则为6列。"""
    first_a = ws.cell(start_row, 1).value
    if first_a is None:
        return False
    try:
        int(first_a)
        return False  # A列是序号 → 5列布局
    except (ValueError, TypeError):
        return True   # A列是分类文字 → 6列布局


def parse_hazard_templates(xlsx_path: str) -> List[Dict]:
    """
    解析隐患依据 Excel，返回模板条目列表。
    自动检测列结构：
      - 5列: 序号(A), 隐患描述(B), 整改建议(C), 依据标准(D), 条款(E)
      - 6列: 类别(A), 序号(B), 隐患描述(C), 整改建议(D), 依据标准(E), 条款(F)
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    templates = []

    all_sheet_names = list(wb.sheetnames)

    for sheet_name in all_sheet_names:
        ws = wb[sheet_name]
        start_row = _detect_start_row(ws)

        if ws.max_row < start_row:
            continue

        is_6col = _detect_col_layout(ws, start_row)

        for row in range(start_row, ws.max_row + 1):
            if is_6col:
                sub_cat = _clean(ws.cell(row, 1).value)
                seq_val = ws.cell(row, 2).value
                desc = _clean(ws.cell(row, 3).value)
                suggestion = _clean(ws.cell(row, 4).value)
                ref_std = _clean(ws.cell(row, 5).value)
                clause = _clean(ws.cell(row, 6).value)
            else:
                sub_cat = ""
                seq_val = ws.cell(row, 1).value
                desc = _clean(ws.cell(row, 2).value)
                suggestion = _clean(ws.cell(row, 3).value)
                ref_std = _clean(ws.cell(row, 4).value)
                clause = _clean(ws.cell(row, 5).value)

            if not desc:
                if ref_std or clause:
                    if templates and templates[-1]["category"] == sheet_name:
                        prev = templates[-1]
                        if ref_std and not prev["reference_standard"]:
                            prev["reference_standard"] = ref_std
                        if clause:
                            prev["standard_clause"] = (
                                (prev["standard_clause"] + "\n" + clause)
                                if prev["standard_clause"] else clause
                            )
                continue

            seq = None
            if seq_val is not None:
                try:
                    seq = int(seq_val)
                except (ValueError, TypeError):
                    pass

            templates.append({
                "category": sheet_name,
                "sub_category": sub_cat if sub_cat and not str(sub_cat).isdigit() else "",
                "seq": seq,
                "description": desc,
                "suggestion": suggestion,
                "reference_standard": ref_std,
                "standard_clause": clause,
            })

    wb.close()

    # 为没有任何模板的 sheet 也保留分类（空分类可通过页面手动添加模板）
    existing_cats = set(t["category"] for t in templates)
    for sn in all_sheet_names:
        if sn not in existing_cats:
            templates.append({
                "category": sn,
                "sub_category": "",
                "seq": 0,
                "description": "（暂无模板，请手动添加）",
                "suggestion": "",
                "reference_standard": "",
                "standard_clause": "",
            })

    return templates


def load_templates_to_db(xlsx_path: str, db_session):
    """解析 Excel 并写入数据库，先清空再导入。"""
    from models import HazardTemplate

    db_session.query(HazardTemplate).delete()

    items = parse_hazard_templates(xlsx_path)
    for item in items:
        t = HazardTemplate(**item)
        db_session.add(t)

    db_session.commit()
    return len(items)
