# -*- coding: utf-8 -*-
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
import os
import shutil
from datetime import datetime

from database import get_db
from models import Project, Hazard, ScenePhoto, GeneratedDoc, HazardTemplate, DetectionRecord, ChecklistResult, Inspector
from config import EXCEL_DIR, HAZARD_PHOTO_DIR, RECTIFY_PHOTO_DIR, SCENE_PHOTO_DIR, DETECTION_PHOTO_DIR
from services.excel_parser import parse_excel
from services.doc_generator import generate_document

VALID_PHOTO_TYPES = ("facade", "card", "license", "permit", "record_sheet")

STREET_GROUP_MAP = {
    "观澜": "A", "观湖": "A", "福城": "A",
    "民治": "B", "龙华": "B", "大浪": "B",
}

DEFAULT_INSPECTORS = {
    "A": "方兵、苏势喆",
    "B": "王宗跃、王金洋",
}


def _resolve_street_group(street: str) -> str:
    """将街道名称映射到分组，支持模糊匹配（如"观澜街道"→"A"）。"""
    if not street:
        return ""
    for key, group in STREET_GROUP_MAP.items():
        if key in street:
            return group
    return ""

VALID_DETECTION_TYPES = (
    "infrared", "ground_resistance", "residual_current", "insulation",
    "terminal", "indoor_wiring", "distribution_box", "ceiling_wiring", "grounding",
)

router = APIRouter(prefix="/api", tags=["projects"])


@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    project_type: str = "longhua",
    db: Session = Depends(get_db),
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "仅支持 .xlsx 格式")
    if project_type not in ("longhua", "wenti"):
        raise HTTPException(400, "project_type 必须为 longhua 或 wenti")

    save_path = os.path.join(EXCEL_DIR, file.filename)
    with open(save_path, "wb") as f:
        content = await file.read()
        f.write(content)

    if project_type == "wenti":
        return _upload_wenti_excel(save_path, file.filename, db)
    else:
        return _upload_longhua_excel(save_path, file.filename, db)


def _upload_longhua_excel(save_path: str, filename: str, db: Session):
    try:
        projects_data = parse_excel(save_path, str(HAZARD_PHOTO_DIR))
    except Exception as e:
        raise HTTPException(500, f"Excel解析失败: {str(e)}")

    created_projects = []
    for pd in projects_data:
        group = _resolve_street_group(pd.street)
        default_insp = DEFAULT_INSPECTORS.get(group, "")
        proj = Project(
            name=pd.name, street=pd.street, address=pd.address,
            contact=pd.contact, phone=pd.phone, category=pd.category,
            build_unit=pd.build_unit, construct_unit=pd.construct_unit,
            supervise_unit=pd.supervise_unit, status="pending",
            excel_file=save_path, project_type="longhua",
            inspectors=default_insp,
        )
        db.add(proj)
        db.flush()

        for hd in pd.hazards:
            hazard = Hazard(
                project_id=proj.id, seq=hd.seq,
                hazard_type=hd.hazard_type, description=hd.description,
                risk=hd.risk, category=hd.category,
                reference=hd.reference, remark=hd.remark,
                hazard_photo_path=hd.hazard_photo_path,
                rectify_status="pending",
            )
            db.add(hazard)

        created_projects.append({
            "id": proj.id, "name": proj.name, "street": proj.street,
            "hazard_count": len(pd.hazards),
            "photo_count": sum(1 for h in pd.hazards if h.hazard_photo_path),
        })

    db.commit()
    return {
        "message": "解析完成",
        "filename": filename,
        "project_count": len(created_projects),
        "projects": created_projects,
    }


def _upload_wenti_excel(save_path: str, filename: str, db: Session):
    from services.excel_parser_wenti import parse_wenti_excel

    try:
        venues = parse_wenti_excel(save_path)
    except Exception as e:
        raise HTTPException(500, f"文体Excel解析失败: {str(e)}")

    created_projects = []
    for v in venues:
        proj = Project(
            name=v.name, street=v.street, address=v.address,
            contact=v.contact, phone=v.phone, category=v.venue_type,
            status="pending", excel_file=save_path,
            project_type="wenti", report_code=v.report_code,
            area=v.area,
        )
        db.add(proj)
        db.flush()
        created_projects.append({
            "id": proj.id, "name": proj.name, "street": proj.street,
            "category": proj.category, "report_code": proj.report_code,
        })

    db.commit()
    return {
        "message": "文体项目解析完成",
        "filename": filename,
        "project_count": len(created_projects),
        "projects": created_projects,
    }


@router.get("/projects")
def list_projects(
    street: Optional[str] = None, status: Optional[str] = None,
    category: Optional[str] = None, project_type: Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(Project)
    if project_type:
        q = q.filter(Project.project_type == project_type)
    if street:
        q = q.filter(Project.street == street)
    if status:
        q = q.filter(Project.status == status)
    if category:
        q = q.filter(Project.category == category)
    projects = q.all()

    result = []
    for p in projects:
        try:
            hazards = db.query(Hazard).filter(Hazard.project_id == p.id).all()
            scene_photos = db.query(ScenePhoto).filter(ScenePhoto.project_id == p.id).all()
            scene_count = len(scene_photos)
            if p.project_type == "wenti":
                hazard_photo_uploaded = sum(1 for h in hazards if h.hazard_photo_path)
                total_needed = 4 + len(hazards)
                total_uploaded = scene_count + hazard_photo_uploaded
                uploaded_count = hazard_photo_uploaded
            else:
                rectify_uploaded = sum(1 for h in hazards if h.rectify_photo_path)
                total_needed = len(hazards) + 2
                total_uploaded = rectify_uploaded + scene_count
                uploaded_count = rectify_uploaded
            excel_name = os.path.basename(p.excel_file).rsplit('.', 1)[0] if p.excel_file else ""
            item = {
                "id": p.id, "name": p.name, "street": p.street,
                "address": p.address, "contact": p.contact, "phone": p.phone,
                "category": p.category, "status": p.status,
                "project_type": p.project_type or "longhua",
                "build_unit": p.build_unit, "construct_unit": p.construct_unit,
                "supervise_unit": p.supervise_unit, "check_date": p.check_date,
                "hazard_count": len(hazards), "uploaded_count": uploaded_count,
                "scene_count": scene_count,
                "total_photos": total_needed,
                "total_uploaded": total_uploaded,
                "source_file": excel_name,
            }
            if p.project_type == "wenti":
                item["report_code"] = p.report_code
                item["area"] = p.area
                item["floor_info"] = p.floor_info
            result.append(item)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise HTTPException(500, f"Error processing project {p.id}: {str(e)}")
    return result


@router.post("/projects")
async def create_project(body: dict, db: Session = Depends(get_db)):
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "项目名称不能为空")

    project_type = body.get("project_type", "wenti")
    if project_type not in ("longhua", "wenti"):
        raise HTTPException(400, "project_type 必须为 longhua 或 wenti")

    proj = Project(
        name=name,
        street=(body.get("street") or "").strip(),
        address=(body.get("address") or "").strip(),
        contact=(body.get("contact") or "").strip(),
        phone=(body.get("phone") or "").strip(),
        category=(body.get("category") or "").strip(),
        status="pending",
        project_type=project_type,
    )

    if project_type == "wenti":
        proj.report_code = (body.get("report_code") or "").strip()
        proj.area = (body.get("area") or "").strip()
        proj.floor_info = (body.get("floor_info") or "").strip()
        proj.inspectors = (body.get("inspectors") or "").strip()
        proj.check_date = (body.get("check_date") or "").strip()
    else:
        proj.build_unit = (body.get("build_unit") or "").strip()
        proj.construct_unit = (body.get("construct_unit") or "").strip()
        proj.supervise_unit = (body.get("supervise_unit") or "").strip()
        group = _resolve_street_group(proj.street)
        proj.inspectors = body.get("inspectors") or DEFAULT_INSPECTORS.get(group, "")
        proj.check_date = (body.get("check_date") or "").strip()

    db.add(proj)
    db.commit()
    db.refresh(proj)

    return {"message": "项目创建成功", "project_id": proj.id, "name": proj.name}


@router.post("/projects/batch-delete")
async def batch_delete_projects(body: dict, db: Session = Depends(get_db)):
    ids = body.get("ids", [])
    if not ids:
        raise HTTPException(400, "请提供要删除的项目ID列表")
    deleted = 0
    for pid in ids:
        proj = db.query(Project).filter(Project.id == pid).first()
        if proj:
            db.delete(proj)
            deleted += 1
    db.commit()
    return {"message": f"已删除 {deleted} 个项目", "deleted": deleted}


@router.get("/projects/{project_id}")
def get_project(project_id: int, db: Session = Depends(get_db)):
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        raise HTTPException(404, "项目不存在")

    hazards = db.query(Hazard).filter(Hazard.project_id == project_id).order_by(Hazard.seq).all()
    scene_photos = db.query(ScenePhoto).filter(ScenePhoto.project_id == project_id).all()

    data = {
        "id": proj.id, "name": proj.name, "street": proj.street,
        "address": proj.address, "contact": proj.contact, "phone": proj.phone,
        "category": proj.category, "status": proj.status,
        "project_type": proj.project_type or "longhua",
        "build_unit": proj.build_unit, "construct_unit": proj.construct_unit,
        "supervise_unit": proj.supervise_unit, "check_date": proj.check_date,
        "inspectors": proj.inspectors or "",
        "hazards": [{
            "id": h.id, "seq": h.seq, "hazard_type": h.hazard_type,
            "description": h.description, "risk": h.risk, "category": h.category,
            "reference": h.reference, "remark": h.remark,
            "suggestion": h.suggestion or "",
            "has_hazard_photo": bool(h.hazard_photo_path) and os.path.exists(h.hazard_photo_path),
            "rectify_status": h.rectify_status,
            "has_rectify_photo": bool(h.rectify_photo_path) and os.path.exists(h.rectify_photo_path),
        } for h in hazards],
        "scene_photos": [{
            "id": sp.id, "photo_type": sp.photo_type,
        } for sp in scene_photos],
    }
    if proj.project_type == "wenti":
        data["report_code"] = proj.report_code
        data["area"] = proj.area
        data["floor_info"] = proj.floor_info
    return data


@router.patch("/projects/{project_id}")
async def update_project(project_id: int, body: dict, db: Session = Depends(get_db)):
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        raise HTTPException(404, "项目不存在")
    updatable = ["name", "street", "address", "contact", "phone", "category",
                 "build_unit", "construct_unit", "supervise_unit", "check_date",
                 "report_code", "area", "floor_info", "inspectors"]
    for field in updatable:
        if field in body:
            setattr(proj, field, body[field])
    db.commit()
    return {"message": "更新成功", "project_id": project_id}


@router.patch("/hazards/{hazard_id}/remark")
async def update_hazard_remark(hazard_id: int, body: dict, db: Session = Depends(get_db)):
    hazard = db.query(Hazard).filter(Hazard.id == hazard_id).first()
    if not hazard:
        raise HTTPException(404, "隐患不存在")
    hazard.remark = body.get("remark", "")
    db.commit()
    return {"message": "备注已更新", "hazard_id": hazard_id, "remark": hazard.remark}


@router.patch("/hazards/{hazard_id}/rectify-status")
async def update_rectify_status(hazard_id: int, body: dict, db: Session = Depends(get_db)):
    hazard = db.query(Hazard).filter(Hazard.id == hazard_id).first()
    if not hazard:
        raise HTTPException(404, "隐患不存在")
    status = body.get("rectify_status", "")
    if status not in ("done", "pending"):
        raise HTTPException(400, "状态值无效，应为 done 或 pending")
    hazard.rectify_status = status
    db.commit()
    _update_project_status(hazard.project_id, db)
    return {"message": "状态已更新", "hazard_id": hazard_id, "rectify_status": status}


@router.post("/hazards/{hazard_id}/rectify-photo")
async def upload_rectify_photo(hazard_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    hazard = db.query(Hazard).filter(Hazard.id == hazard_id).first()
    if not hazard:
        raise HTTPException(404, "隐患不存在")

    ext = os.path.splitext(file.filename)[1] or ".jpg"
    filename = f"rectify_p{hazard.project_id}_h{hazard_id}{ext}"
    save_path = os.path.join(RECTIFY_PHOTO_DIR, filename)

    with open(save_path, "wb") as f:
        content = await file.read()
        f.write(content)

    hazard.rectify_photo_path = save_path
    hazard.rectify_status = "done"
    db.commit()

    _update_project_status(hazard.project_id, db)

    return {"message": "上传成功", "hazard_id": hazard_id}


@router.post("/projects/{project_id}/scene-photo")
async def upload_scene_photo(
    project_id: int, photo_type: str, file: UploadFile = File(...), db: Session = Depends(get_db)
):
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        raise HTTPException(404, "项目不存在")
    if photo_type not in VALID_PHOTO_TYPES:
        raise HTTPException(400, f"photo_type 必须为 {', '.join(VALID_PHOTO_TYPES)} 之一")

    ext = os.path.splitext(file.filename)[1] or ".jpg"
    filename = f"scene_p{project_id}_{photo_type}{ext}"
    save_path = os.path.join(SCENE_PHOTO_DIR, filename)

    with open(save_path, "wb") as f:
        content = await file.read()
        f.write(content)

    existing = db.query(ScenePhoto).filter(
        ScenePhoto.project_id == project_id, ScenePhoto.photo_type == photo_type
    ).first()
    if existing:
        existing.photo_path = save_path
    else:
        sp = ScenePhoto(project_id=project_id, photo_type=photo_type, photo_path=save_path)
        db.add(sp)

    db.commit()
    _update_project_status(project_id, db)

    return {"message": "上传成功", "photo_type": photo_type}


@router.get("/projects/{project_id}/scene-photo")
def get_scene_photo(project_id: int, photo_type: str, db: Session = Depends(get_db)):
    from fastapi.responses import FileResponse

    if photo_type not in VALID_PHOTO_TYPES:
        raise HTTPException(400, f"photo_type 必须为 {', '.join(VALID_PHOTO_TYPES)} 之一")
    sp = (
        db.query(ScenePhoto)
        .filter(ScenePhoto.project_id == project_id, ScenePhoto.photo_type == photo_type)
        .first()
    )
    if not sp or not sp.photo_path or not os.path.exists(sp.photo_path):
        raise HTTPException(404, "现场照片不存在")
    return FileResponse(sp.photo_path)


@router.post("/projects/{project_id}/generate")
def generate_doc(project_id: int, db: Session = Depends(get_db)):
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        raise HTTPException(404, "项目不存在")

    hazards = db.query(Hazard).filter(Hazard.project_id == project_id).order_by(Hazard.seq).all()
    scene_photos = db.query(ScenePhoto).filter(ScenePhoto.project_id == project_id).all()

    if proj.project_type == "wenti":
        output_path = _generate_wenti(proj, hazards, scene_photos, db)
    else:
        output_path = _generate_longhua(proj, hazards, scene_photos)

    file_size = os.path.getsize(output_path)
    file_name = os.path.basename(output_path)

    existing_doc = db.query(GeneratedDoc).filter(GeneratedDoc.project_id == project_id).first()
    if existing_doc:
        if existing_doc.file_path != output_path and os.path.exists(existing_doc.file_path):
            try:
                os.remove(existing_doc.file_path)
            except OSError:
                pass
        existing_doc.file_path = output_path
        existing_doc.file_name = file_name
        existing_doc.file_size = file_size
        existing_doc.created_at = datetime.now()
    else:
        gen_doc = GeneratedDoc(
            project_id=project_id, file_path=output_path,
            file_name=file_name, file_size=file_size,
        )
        db.add(gen_doc)

    db.commit()

    return {
        "message": "生成成功",
        "file_name": file_name,
        "file_size": file_size,
        "download_url": f"/api/projects/{project_id}/download",
    }


def _generate_longhua(proj, hazards, scene_photos) -> str:
    facade_photo = next((sp.photo_path for sp in scene_photos if sp.photo_type == "facade"), None)
    card_photo = next((sp.photo_path for sp in scene_photos if sp.photo_type == "card"), None)

    proj_dict = {
        "name": proj.name, "street": proj.street, "address": proj.address,
        "inspectors": proj.inspectors or "",
        "check_date": proj.check_date or "",
    }
    hz_list = [{
        "seq": h.seq, "hazard_type": h.hazard_type,
        "description": h.description, "risk": h.risk,
        "hazard_photo_path": h.hazard_photo_path,
        "rectify_status": "已整改" if h.rectify_status == "done" else "未整改",
        "rectify_photo_path": h.rectify_photo_path,
        "remark": h.remark or "",
    } for h in hazards]

    try:
        return generate_document(proj_dict, hz_list, facade_photo, card_photo)
    except Exception as e:
        raise HTTPException(500, f"文档生成失败: {str(e)}")


def _generate_wenti(proj, hazards, scene_photos, db=None) -> str:
    from services.doc_generator_wenti import generate_wenti_document

    proj_dict = {
        "name": proj.name, "street": proj.street, "address": proj.address,
        "contact": proj.contact, "phone": proj.phone, "category": proj.category,
        "check_date": proj.check_date or "",
        "report_code": proj.report_code or "",
        "area": proj.area or "",
        "floor_info": proj.floor_info or "",
        "inspectors": proj.inspectors or "",
    }
    hz_list = [{
        "seq": h.seq, "hazard_type": h.hazard_type,
        "description": h.description,
        "suggestion": h.suggestion or "",
        "hazard_photo_path": h.hazard_photo_path,
        "reference": h.reference or "",
    } for h in hazards]

    photos = {}
    for sp in scene_photos:
        if sp.photo_path and os.path.exists(sp.photo_path):
            photos[sp.photo_type] = sp.photo_path

    detection_records = []
    checklist_results = []
    if db:
        det_rows = db.query(DetectionRecord).filter(
            DetectionRecord.project_id == proj.id
        ).order_by(DetectionRecord.detection_type, DetectionRecord.seq).all()
        detection_records = [{
            "id": r.id, "detection_type": r.detection_type, "seq": r.seq,
            "location": r.location or "", "photo_path": r.photo_path or "",
            "code": r.code or "", "temperature": r.temperature or "",
            "resistance_value": r.resistance_value or "",
            "result": r.result or "", "remark": r.remark or "",
        } for r in det_rows]

        cl_rows = db.query(ChecklistResult).filter(
            ChecklistResult.project_id == proj.id
        ).order_by(ChecklistResult.table_index, ChecklistResult.item_seq).all()
        checklist_results = [{
            "table_index": c.table_index, "item_seq": c.item_seq, "result": c.result or "",
        } for c in cl_rows]

    try:
        return generate_wenti_document(proj_dict, hz_list, photos, detection_records, checklist_results)
    except Exception as e:
        raise HTTPException(500, f"文体文档生成失败: {str(e)}")


@router.get("/projects/{project_id}/download")
def download_doc(project_id: int, db: Session = Depends(get_db)):
    from fastapi.responses import FileResponse

    gen_doc = db.query(GeneratedDoc).filter(GeneratedDoc.project_id == project_id).first()
    if not gen_doc or not os.path.exists(gen_doc.file_path):
        raise HTTPException(404, "文档不存在，请先生成")

    return FileResponse(
        gen_doc.file_path,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        filename=gen_doc.file_name,
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


@router.get("/projects/{project_id}/hazard-photo/{hazard_id}")
def get_hazard_photo(project_id: int, hazard_id: int, db: Session = Depends(get_db)):
    from fastapi.responses import FileResponse

    hazard = db.query(Hazard).filter(Hazard.id == hazard_id, Hazard.project_id == project_id).first()
    if not hazard or not hazard.hazard_photo_path or not os.path.exists(hazard.hazard_photo_path):
        raise HTTPException(404, "隐患照片不存在")

    return FileResponse(hazard.hazard_photo_path)


@router.get("/projects/{project_id}/rectify-photo/{hazard_id}")
def get_rectify_photo(project_id: int, hazard_id: int, db: Session = Depends(get_db)):
    from fastapi.responses import FileResponse

    hazard = db.query(Hazard).filter(Hazard.id == hazard_id, Hazard.project_id == project_id).first()
    if not hazard or not hazard.rectify_photo_path or not os.path.exists(hazard.rectify_photo_path):
        raise HTTPException(404, "整改照片不存在")

    return FileResponse(hazard.rectify_photo_path)


@router.get("/generated-docs")
def list_generated_docs(db: Session = Depends(get_db)):
    docs = db.query(GeneratedDoc).all()
    result = []
    for d in docs:
        proj = db.query(Project).filter(Project.id == d.project_id).first()
        result.append({
            "id": d.id, "project_id": d.project_id,
            "project_name": proj.name if proj else "",
            "file_name": d.file_name, "file_size": d.file_size,
            "created_at": d.created_at.isoformat() if d.created_at else "",
            "download_url": f"/api/projects/{d.project_id}/download",
        })
    return result


@router.get("/streets")
def list_streets(db: Session = Depends(get_db)):
    streets = db.query(Project.street).distinct().all()
    return [s[0] for s in streets if s[0]]


@router.get("/categories")
def list_categories(db: Session = Depends(get_db)):
    cats = db.query(Project.category).distinct().all()
    return [c[0] for c in cats if c[0]]


@router.delete("/projects/{project_id}")
def delete_project(project_id: int, db: Session = Depends(get_db)):
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        raise HTTPException(404, "项目不存在")
    db.delete(proj)
    db.commit()
    return {"message": "删除成功", "project_id": project_id}


@router.post("/projects/{project_id}/re-extract-photos")
def re_extract_photos(project_id: int, db: Session = Depends(get_db)):
    """Re-extract hazard photos from the original Excel for a single project."""
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        raise HTTPException(404, "项目不存在")
    if not proj.excel_file or not os.path.exists(proj.excel_file):
        raise HTTPException(400, "原始Excel文件不存在，无法重新提取")

    from services.excel_parser import _extract_images_from_xlsx, _find_photo_for_row
    image_map = _extract_images_from_xlsx(proj.excel_file, str(HAZARD_PHOTO_DIR))

    hazards = db.query(Hazard).filter(Hazard.project_id == project_id).order_by(Hazard.seq).all()

    import openpyxl
    wb = openpyxl.load_workbook(proj.excel_file, data_only=True)
    ws = wb.active
    row_map = _build_hazard_row_map(ws, proj.name)
    wb.close()

    updated = 0
    for h in hazards:
        if h.hazard_photo_path and os.path.exists(h.hazard_photo_path):
            continue
        anchor_row = row_map.get(h.seq)
        if anchor_row is None:
            continue
        photo_path = _find_photo_for_row(image_map, anchor_row)
        if photo_path:
            h.hazard_photo_path = photo_path
            updated += 1

    db.commit()
    return {
        "message": f"重新提取完成，更新了 {updated} 条隐患照片",
        "project_id": project_id,
        "updated": updated,
        "total_hazards": len(hazards),
    }


@router.post("/re-extract-all-photos")
def re_extract_all_photos(db: Session = Depends(get_db)):
    """Re-extract hazard photos for ALL projects that have missing photos."""
    projects = db.query(Project).all()
    total_updated = 0
    project_results = []

    excel_cache: dict = {}

    for proj in projects:
        if not proj.excel_file or not os.path.exists(proj.excel_file):
            continue

        hazards = db.query(Hazard).filter(Hazard.project_id == proj.id).order_by(Hazard.seq).all()
        missing = [h for h in hazards if not h.hazard_photo_path or not os.path.exists(h.hazard_photo_path or "")]
        if not missing:
            continue

        if proj.excel_file not in excel_cache:
            from services.excel_parser import _extract_images_from_xlsx
            image_map = _extract_images_from_xlsx(proj.excel_file, str(HAZARD_PHOTO_DIR))
            import openpyxl
            wb = openpyxl.load_workbook(proj.excel_file, data_only=True)
            ws = wb.active
            excel_cache[proj.excel_file] = (image_map, ws, wb)
        else:
            image_map, ws, wb = excel_cache[proj.excel_file]

        row_map = _build_hazard_row_map(ws, proj.name)
        from services.excel_parser import _find_photo_for_row

        updated = 0
        for h in missing:
            anchor_row = row_map.get(h.seq)
            if anchor_row is None:
                continue
            photo_path = _find_photo_for_row(image_map, anchor_row)
            if photo_path:
                h.hazard_photo_path = photo_path
                updated += 1

        if updated:
            total_updated += updated
            project_results.append({"id": proj.id, "name": proj.name, "updated": updated})

    for _, _, wb in excel_cache.values():
        wb.close()

    db.commit()
    return {
        "message": f"批量重新提取完成，共更新 {total_updated} 条隐患照片",
        "total_updated": total_updated,
        "projects": project_results,
    }


def _build_hazard_row_map(ws, project_name: str) -> dict:
    """
    Scan the worksheet to find which Excel rows correspond to hazard seq numbers
    for the given project. Returns {seq: 0-based-row}.
    """
    import re
    from services.excel_parser import _clean

    max_row = ws.max_row
    row_map = {}
    found_project = False
    row = 1

    while row <= max_row:
        c1 = _clean(ws.cell(row, 1).value)
        if not c1:
            row += 1
            continue

        if re.search(r"项目名称[：:]", c1):
            name_match = re.search(r"项目名称[：:]\s*(.+?)(?:\n|$)", c1.replace("\r\n", "\n"))
            if name_match and name_match.group(1).strip() == project_name:
                found_project = True
                row += 1
                if row <= max_row and _clean(ws.cell(row, 1).value) == "序号":
                    row += 1
                while row <= max_row:
                    seq_val = ws.cell(row, 1).value
                    if seq_val is None:
                        row += 1
                        continue
                    seq_str = _clean(seq_val)
                    if not seq_str.isdigit():
                        break
                    row_map[int(seq_str)] = row - 1  # 0-based
                    row += 1
                break
            else:
                row += 1
                continue

        row += 1

    return row_map


def _update_project_status(project_id: int, db: Session):
    """Update project status based on photo upload progress."""
    hazards = db.query(Hazard).filter(Hazard.project_id == project_id).all()
    scene_photos = db.query(ScenePhoto).filter(ScenePhoto.project_id == project_id).all()

    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        return

    any_uploaded = any(h.rectify_photo_path for h in hazards) or len(scene_photos) > 0
    has_facade = any(sp.photo_type == "facade" for sp in scene_photos)

    if proj.project_type == "wenti":
        hazard_photos_done = all(h.hazard_photo_path for h in hazards) if hazards else True
        any_hazard_photo = any(h.hazard_photo_path for h in hazards)
        if hazard_photos_done and (len(hazards) > 0 or has_facade):
            proj.status = "done"
        elif any_hazard_photo or any_uploaded or len(hazards) > 0:
            proj.status = "progress"
    else:
        all_rectified = all(h.rectify_photo_path for h in hazards)
        has_card = any(sp.photo_type == "card" for sp in scene_photos)
        if all_rectified and has_facade and has_card:
            proj.status = "done"
        elif any_uploaded:
            proj.status = "progress"

    db.commit()


# ──────────────────── 隐患模板 API ────────────────────

@router.get("/hazard-templates")
def list_hazard_templates(category: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(HazardTemplate)
    if category:
        q = q.filter(HazardTemplate.category == category)
    templates = q.order_by(HazardTemplate.category, HazardTemplate.seq).all()
    return [{
        "id": t.id, "category": t.category, "sub_category": t.sub_category,
        "seq": t.seq, "description": t.description, "suggestion": t.suggestion,
        "reference_standard": t.reference_standard, "standard_clause": t.standard_clause,
    } for t in templates]


@router.get("/hazard-template-categories")
def list_template_categories(db: Session = Depends(get_db)):
    cats = db.query(HazardTemplate.category).distinct().all()
    return [c[0] for c in cats if c[0]]


@router.post("/hazard-templates/load")
async def load_hazard_templates(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "仅支持 .xlsx 格式")

    save_path = os.path.join(EXCEL_DIR, "hazard_templates_" + file.filename)
    with open(save_path, "wb") as f:
        content = await file.read()
        f.write(content)

    from services.template_loader import load_templates_to_db
    try:
        count = load_templates_to_db(save_path, db)
    except Exception as e:
        raise HTTPException(500, f"模板导入失败: {str(e)}")

    return {"message": f"已导入 {count} 条隐患模板", "count": count}


@router.post("/hazard-templates")
def create_hazard_template(body: dict, db: Session = Depends(get_db)):
    max_seq = db.query(HazardTemplate.seq).filter(
        HazardTemplate.category == body.get("category", "")
    ).order_by(HazardTemplate.seq.desc()).first()
    next_seq = (max_seq[0] + 1) if max_seq and max_seq[0] else 1

    t = HazardTemplate(
        category=body.get("category", ""),
        sub_category=body.get("sub_category", ""),
        seq=body.get("seq") or next_seq,
        description=body.get("description", ""),
        suggestion=body.get("suggestion", ""),
        reference_standard=body.get("reference_standard", ""),
        standard_clause=body.get("standard_clause", ""),
    )
    db.add(t)
    db.commit()
    db.refresh(t)
    return {"id": t.id, "message": "模板已添加"}


@router.put("/hazard-templates/{template_id}")
def update_hazard_template(template_id: int, body: dict, db: Session = Depends(get_db)):
    t = db.query(HazardTemplate).filter(HazardTemplate.id == template_id).first()
    if not t:
        raise HTTPException(404, "模板不存在")
    for field in ("category", "sub_category", "seq", "description", "suggestion",
                  "reference_standard", "standard_clause"):
        if field in body:
            setattr(t, field, body[field])
    db.commit()
    return {"message": "模板已更新"}


@router.delete("/hazard-templates/{template_id}")
def delete_hazard_template(template_id: int, db: Session = Depends(get_db)):
    t = db.query(HazardTemplate).filter(HazardTemplate.id == template_id).first()
    if not t:
        raise HTTPException(404, "模板不存在")
    db.delete(t)
    db.commit()
    return {"message": "模板已删除"}


# ──────────────────── 隐患 CRUD API ────────────────────

@router.post("/projects/{project_id}/hazards")
async def add_hazard(project_id: int, body: dict, db: Session = Depends(get_db)):
    """添加隐患条目到项目（可从模板或手动录入）。"""
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        raise HTTPException(404, "项目不存在")

    max_seq = db.query(Hazard.seq).filter(Hazard.project_id == project_id).order_by(Hazard.seq.desc()).first()
    next_seq = (max_seq[0] + 1) if max_seq and max_seq[0] else 1

    hazard = Hazard(
        project_id=project_id,
        seq=body.get("seq", next_seq),
        hazard_type=body.get("hazard_type", ""),
        description=body.get("description", ""),
        suggestion=body.get("suggestion", ""),
        risk=body.get("risk", ""),
        category=body.get("category", ""),
        reference=body.get("reference", ""),
        remark=body.get("remark", ""),
        rectify_status="pending",
    )
    db.add(hazard)
    db.commit()
    db.refresh(hazard)

    return {"message": "隐患已添加", "hazard_id": hazard.id, "seq": hazard.seq}


@router.patch("/hazards/{hazard_id}")
async def update_hazard(hazard_id: int, body: dict, db: Session = Depends(get_db)):
    """编辑隐患条目。"""
    hazard = db.query(Hazard).filter(Hazard.id == hazard_id).first()
    if not hazard:
        raise HTTPException(404, "隐患不存在")

    updatable = ["hazard_type", "description", "suggestion", "risk",
                 "category", "reference", "remark", "seq"]
    for field in updatable:
        if field in body:
            setattr(hazard, field, body[field])
    db.commit()
    return {"message": "隐患已更新", "hazard_id": hazard_id}


@router.delete("/hazards/{hazard_id}")
def delete_hazard(hazard_id: int, db: Session = Depends(get_db)):
    hazard = db.query(Hazard).filter(Hazard.id == hazard_id).first()
    if not hazard:
        raise HTTPException(404, "隐患不存在")
    project_id = hazard.project_id
    db.delete(hazard)
    db.commit()
    return {"message": "隐患已删除", "hazard_id": hazard_id, "project_id": project_id}


@router.post("/hazards/{hazard_id}/photo")
async def upload_hazard_photo(hazard_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """上传隐患照片（文体项目用）。"""
    hazard = db.query(Hazard).filter(Hazard.id == hazard_id).first()
    if not hazard:
        raise HTTPException(404, "隐患不存在")

    ext = os.path.splitext(file.filename)[1] or ".jpg"
    filename = f"hazard_p{hazard.project_id}_h{hazard_id}{ext}"
    save_path = os.path.join(HAZARD_PHOTO_DIR, filename)

    with open(save_path, "wb") as f:
        content = await file.read()
        f.write(content)

    hazard.hazard_photo_path = save_path
    db.commit()

    _update_project_status(hazard.project_id, db)

    return {"message": "隐患照片上传成功", "hazard_id": hazard_id}


# ──────────────────── 检测记录 CRUD API ────────────────────

@router.get("/projects/{project_id}/detections")
def list_detections(project_id: int, detection_type: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(DetectionRecord).filter(DetectionRecord.project_id == project_id)
    if detection_type:
        q = q.filter(DetectionRecord.detection_type == detection_type)
    records = q.order_by(DetectionRecord.detection_type, DetectionRecord.seq).all()
    return [{
        "id": r.id, "project_id": r.project_id,
        "detection_type": r.detection_type, "seq": r.seq,
        "location": r.location or "", "photo_path": r.photo_path or "",
        "has_photo": bool(r.photo_path) and os.path.exists(r.photo_path or ""),
        "code": r.code or "", "temperature": r.temperature or "",
        "resistance_value": r.resistance_value or "",
        "result": r.result or "", "remark": r.remark or "",
    } for r in records]


@router.post("/projects/{project_id}/detections")
async def add_detection(project_id: int, body: dict, db: Session = Depends(get_db)):
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        raise HTTPException(404, "项目不存在")

    det_type = body.get("detection_type", "")
    if det_type not in VALID_DETECTION_TYPES:
        raise HTTPException(400, f"detection_type 必须为 {', '.join(VALID_DETECTION_TYPES)} 之一")

    max_seq = db.query(DetectionRecord.seq).filter(
        DetectionRecord.project_id == project_id,
        DetectionRecord.detection_type == det_type,
    ).order_by(DetectionRecord.seq.desc()).first()
    next_seq = (max_seq[0] + 1) if max_seq and max_seq[0] else 1

    result_val = body.get("result", "")
    if det_type == "infrared" and not result_val:
        try:
            temp = float(body.get("temperature", "0") or "0")
            result_val = "不符合规范要求" if temp > 70 else "符合规范要求"
        except (ValueError, TypeError):
            pass
    if det_type == "ground_resistance" and not result_val:
        try:
            res = float(body.get("resistance_value", "0") or "0")
            result_val = "不符合规范要求" if res > 4 else "符合规范要求"
        except (ValueError, TypeError):
            pass

    record = DetectionRecord(
        project_id=project_id,
        detection_type=det_type,
        seq=body.get("seq", next_seq),
        location=body.get("location", ""),
        code=body.get("code", ""),
        temperature=body.get("temperature", ""),
        resistance_value=body.get("resistance_value", ""),
        result=result_val,
        remark=body.get("remark", ""),
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return {"message": "检测记录已添加", "id": record.id, "seq": record.seq, "result": record.result}


@router.patch("/detections/{record_id}")
async def update_detection(record_id: int, body: dict, db: Session = Depends(get_db)):
    record = db.query(DetectionRecord).filter(DetectionRecord.id == record_id).first()
    if not record:
        raise HTTPException(404, "检测记录不存在")

    updatable = ["location", "code", "temperature", "resistance_value", "result", "remark"]
    for field in updatable:
        if field in body:
            setattr(record, field, body[field])

    if record.detection_type == "infrared" and "temperature" in body and "result" not in body:
        try:
            temp = float(body["temperature"] or "0")
            record.result = "不符合规范要求" if temp > 70 else "符合规范要求"
        except (ValueError, TypeError):
            pass
    if record.detection_type == "ground_resistance" and "resistance_value" in body and "result" not in body:
        try:
            res = float(body["resistance_value"] or "0")
            record.result = "不符合规范要求" if res > 4 else "符合规范要求"
        except (ValueError, TypeError):
            pass

    db.commit()
    return {"message": "检测记录已更新", "id": record_id, "result": record.result}


@router.delete("/detections/{record_id}")
def delete_detection(record_id: int, db: Session = Depends(get_db)):
    record = db.query(DetectionRecord).filter(DetectionRecord.id == record_id).first()
    if not record:
        raise HTTPException(404, "检测记录不存在")
    db.delete(record)
    db.commit()
    return {"message": "检测记录已删除", "id": record_id}


@router.post("/detections/{record_id}/photo")
async def upload_detection_photo(record_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    record = db.query(DetectionRecord).filter(DetectionRecord.id == record_id).first()
    if not record:
        raise HTTPException(404, "检测记录不存在")

    ext = os.path.splitext(file.filename)[1] or ".jpg"
    filename = f"det_p{record.project_id}_{record.detection_type}_{record_id}{ext}"
    save_path = os.path.join(DETECTION_PHOTO_DIR, filename)

    with open(save_path, "wb") as f:
        content = await file.read()
        f.write(content)

    record.photo_path = save_path
    db.commit()
    return {"message": "检测照片上传成功", "id": record_id}


@router.get("/detections/{record_id}/photo")
def get_detection_photo(record_id: int, db: Session = Depends(get_db)):
    from fastapi.responses import FileResponse
    record = db.query(DetectionRecord).filter(DetectionRecord.id == record_id).first()
    if not record or not record.photo_path or not os.path.exists(record.photo_path):
        raise HTTPException(404, "检测照片不存在")
    return FileResponse(record.photo_path)


# ──────────────────── 附表检测结果 API ────────────────────

@router.get("/projects/{project_id}/checklist")
def list_checklist(project_id: int, table_index: Optional[int] = None, db: Session = Depends(get_db)):
    from services.checklist_defaults import CHECKLIST_ITEM_TEXTS
    q = db.query(ChecklistResult).filter(ChecklistResult.project_id == project_id)
    if table_index is not None:
        q = q.filter(ChecklistResult.table_index == table_index)
    results = q.order_by(ChecklistResult.table_index, ChecklistResult.item_seq).all()
    return [{
        "id": r.id, "table_index": r.table_index,
        "item_seq": r.item_seq, "result": r.result or "",
        "text": CHECKLIST_ITEM_TEXTS.get((r.table_index, r.item_seq), ""),
    } for r in results]


@router.patch("/projects/{project_id}/checklist")
async def update_checklist(project_id: int, body: dict, db: Session = Depends(get_db)):
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        raise HTTPException(404, "项目不存在")

    items = body.get("items", [])
    updated = 0
    for item in items:
        ti = item.get("table_index")
        seq = item.get("item_seq")
        result = item.get("result", "")
        if ti is None or seq is None:
            continue
        existing = db.query(ChecklistResult).filter(
            ChecklistResult.project_id == project_id,
            ChecklistResult.table_index == ti,
            ChecklistResult.item_seq == seq,
        ).first()
        if existing:
            existing.result = result
        else:
            db.add(ChecklistResult(
                project_id=project_id, table_index=ti,
                item_seq=seq, result=result,
            ))
        updated += 1

    db.commit()
    return {"message": f"已更新 {updated} 条检测结果", "updated": updated}


@router.post("/projects/{project_id}/init-checklist")
def init_checklist(project_id: int, db: Session = Depends(get_db)):
    """从模板初始化附表检测结果（Tables 17-24 的默认值）。"""
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        raise HTTPException(404, "项目不存在")

    existing = db.query(ChecklistResult).filter(ChecklistResult.project_id == project_id).count()
    if existing > 0:
        return {"message": "附表结果已存在，跳过初始化", "count": existing}

    from services.checklist_defaults import CHECKLIST_DEFAULTS
    count = 0
    for table_idx, items in CHECKLIST_DEFAULTS.items():
        for seq, result in items:
            db.add(ChecklistResult(
                project_id=project_id, table_index=table_idx,
                item_seq=seq, result=result,
            ))
            count += 1
    db.commit()
    return {"message": f"已初始化 {count} 条附表检测结果", "count": count}


@router.get("/inspectors")
def list_inspectors(street: Optional[str] = None, db: Session = Depends(get_db)):
    """按街道筛选核销人员列表。不传 street 则返回全部。"""
    q = db.query(Inspector)
    if street:
        group = _resolve_street_group(street)
        if group:
            q = q.filter(Inspector.street_group == group)
    rows = q.order_by(Inspector.id).all()
    return [{"id": r.id, "name": r.name, "street_group": r.street_group} for r in rows]


@router.post("/inspectors")
def add_inspector(body: dict, db: Session = Depends(get_db)):
    """添加核销人员。body: { name, street } 或 { name, street_group }。"""
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "name 不能为空")
    group = (body.get("street_group") or "").strip()
    if not group:
        street = (body.get("street") or "").strip()
        group = _resolve_street_group(street)
    if group not in ("A", "B"):
        raise HTTPException(400, "无法识别对应的分组，请传入 street 或 street_group(A/B)")
    existing = db.query(Inspector).filter(Inspector.name == name, Inspector.street_group == group).first()
    if existing:
        return {"id": existing.id, "name": existing.name, "street_group": existing.street_group}
    ins = Inspector(name=name, street_group=group)
    db.add(ins)
    db.commit()
    db.refresh(ins)
    return {"id": ins.id, "name": ins.name, "street_group": ins.street_group}


@router.delete("/inspectors/{inspector_id}")
def delete_inspector(inspector_id: int, db: Session = Depends(get_db)):
    ins = db.query(Inspector).filter(Inspector.id == inspector_id).first()
    if not ins:
        raise HTTPException(404, "人员不存在")
    db.delete(ins)
    db.commit()
    return {"message": "删除成功"}
