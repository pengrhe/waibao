# -*- coding: utf-8 -*-
"""
文体项目 Excel 解析器：解析扁平场所列表台账
列：序号(A) 街道(B) 场所类型(C) 场所名称(D) 地址(E) 负责人(F) 联系方式(G) 面积(H) ... 报告编号(M)
"""
import re
import openpyxl
from typing import List
from dataclasses import dataclass


@dataclass
class WentiProjectData:
    seq: int
    street: str
    venue_type: str
    name: str
    address: str
    contact: str
    phone: str
    area: str
    report_code: str


def _clean(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _find_report_code_col(ws) -> int:
    """扫描表头行（第1-2行），查找'文件编号'或'报告编号'所在列。"""
    for r in (2, 1):
        for c in range(1, ws.max_column + 1):
            val = _clean(ws.cell(r, c).value)
            if val in ("文件编号", "报告编号", "编号", "文件名称"):
                return c
    return ws.max_column  # fallback: 最后一列


def parse_wenti_excel(xlsx_path: str) -> List[WentiProjectData]:
    """
    解析文体项目台账 Excel。
    第1行为标题，第2行为表头，第3行起为数据。
    自动检测报告编号所在列。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    max_row = ws.max_row

    code_col = _find_report_code_col(ws)

    projects: List[WentiProjectData] = []

    for row in range(3, max_row + 1):
        seq_val = ws.cell(row, 1).value
        if seq_val is None:
            continue
        try:
            seq = int(seq_val)
        except (ValueError, TypeError):
            continue

        name = _clean(ws.cell(row, 4).value)
        if not name:
            continue

        raw_code = _clean(ws.cell(row, code_col).value).replace("\n", "")
        report_code = raw_code

        proj = WentiProjectData(
            seq=seq,
            street=_clean(ws.cell(row, 2).value),
            venue_type=_clean(ws.cell(row, 3).value),
            name=name,
            address=_clean(ws.cell(row, 5).value),
            contact=_clean(ws.cell(row, 6).value),
            phone=_clean(ws.cell(row, 7).value),
            area=_clean(ws.cell(row, 8).value),
            report_code=report_code,
        )
        projects.append(proj)

    wb.close()
    return projects


if __name__ == "__main__":
    import sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    xlsx = r"d:\wp\waibao\gongcheng\文体\（B包）2025-2026文体旅游行业安全隐患排查服务项目隐患台账（第二轮）(1).xlsx"
    projects = parse_wenti_excel(xlsx)
    print(f"解析完成：共 {len(projects)} 个场所")
    for p in projects[:5]:
        print(f"  [{p.street}] {p.name} - {p.venue_type} | 编号: {p.report_code}")
